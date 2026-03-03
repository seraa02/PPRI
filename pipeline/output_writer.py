"""
pipeline/output_writer.py
--------------------------
Generates the final Excel workbook and CSV files from all processed records.

Excel layout (one sheet per survey type):
  Row 1:  Study Name — merged across all data columns
  Row 2:  "PRE Survey Results" or "POST Survey Results" — merged
  Row 3:  Column headers:
            Participant_ID | Name | Date | Q1 | Q2 | ... | Flagged_Questions
  Row 4+: One row per participant

A flat CSV version of each sheet is also exported.

Output files are saved to OUTPUT_FOLDER:
  {STUDY_NAME}_results.xlsx
  {STUDY_NAME}_PRE_results.csv
  {STUDY_NAME}_POST_results.csv
"""

import csv
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_FOLDER, STUDY_NAME

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_STUDY_ROW_FILL   = PatternFill("solid", fgColor="1F4E79")   # Dark blue
_LABEL_ROW_FILL   = PatternFill("solid", fgColor="2E75B6")   # Medium blue
_HEADER_ROW_FILL  = PatternFill("solid", fgColor="BDD7EE")   # Light blue
_FLAGGED_FILL     = PatternFill("solid", fgColor="FFE699")   # Yellow
_WHITE_FONT       = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
_HEADER_FONT      = Font(name="Calibri", bold=True, color="1F4E79", size=11)
_CELL_FONT        = Font(name="Calibri", size=10)
_THIN_BORDER      = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sort_key(label: str) -> tuple:
    """Natural sort key for question labels: Q1 < Q2 < Q10."""
    digits = "".join(filter(str.isdigit, label))
    prefix = "".join(filter(str.isalpha, label)).upper()
    return (prefix, int(digits) if digits else 0, label)


def _collect_question_columns(records: List[dict]) -> List[str]:
    """
    Return a sorted list of all unique question numbers found across all records.
    """
    seen: set = set()
    for rec in records:
        seen.update(rec.get("questions", {}).keys())
    return sorted(seen, key=_sort_key)


def _participant_id(record: dict) -> str:
    """Derive a Participant_ID from filename (strip extension)."""
    return Path(record.get("filename", "unknown")).stem


def _apply_cell_style(
    cell,
    font=None,
    fill=None,
    alignment=None,
    border=None,
) -> None:
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def _write_sheet(
    wb: Workbook,
    sheet_title: str,
    records: List[dict],
    question_cols: List[str],
    survey_label: str,
) -> Optional[pd.DataFrame]:
    """
    Write one survey-type sheet into the workbook.

    Returns a pandas DataFrame of the data rows (for CSV export), or None
    if there are no records.
    """
    if not records:
        logger.warning(f"No records for {survey_label} — sheet not written.")
        return None

    ws = wb.create_sheet(title=sheet_title)

    # Total number of data columns
    all_headers = ["Participant_ID", "Name", "Date"] + question_cols + ["Flagged_Questions"]
    n_cols = len(all_headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ------------------------------------------------------------------
    # Row 1 — Study name (merged)
    # ------------------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=STUDY_NAME)
    _apply_cell_style(cell, font=_WHITE_FONT, fill=_STUDY_ROW_FILL,
                      alignment=center, border=_THIN_BORDER)
    ws.row_dimensions[1].height = 24

    # ------------------------------------------------------------------
    # Row 2 — Survey type label (merged)
    # ------------------------------------------------------------------
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    cell = ws.cell(row=2, column=1, value=survey_label)
    _apply_cell_style(cell, font=_WHITE_FONT, fill=_LABEL_ROW_FILL,
                      alignment=center, border=_THIN_BORDER)
    ws.row_dimensions[2].height = 20

    # ------------------------------------------------------------------
    # Row 3 — Column headers
    # ------------------------------------------------------------------
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        _apply_cell_style(cell, font=_HEADER_FONT, fill=_HEADER_ROW_FILL,
                          alignment=center, border=_THIN_BORDER)
    ws.row_dimensions[3].height = 18

    # ------------------------------------------------------------------
    # Rows 4+ — Data
    # ------------------------------------------------------------------
    csv_rows: List[dict] = []

    for row_offset, rec in enumerate(records):
        excel_row = 4 + row_offset
        participant_id = _participant_id(rec)
        pi = rec.get("personal_info", {})
        questions = rec.get("questions", {})
        flagged = rec.get("flagged_questions", [])
        is_flagged_row = bool(flagged)

        row_data: List[Any] = [
            participant_id,
            pi.get("name", "UNCLEAR"),
            pi.get("date", "UNCLEAR"),
        ]

        for q_num in question_cols:
            q_info = questions.get(q_num, {})
            answer = q_info.get("answer", "")
            # Convert list answers to a readable string
            if isinstance(answer, list):
                answer = "; ".join(str(a) for a in answer)
            row_data.append(answer if answer != "" else "")

        flagged_str = ", ".join(flagged) if flagged else ""
        row_data.append(flagged_str)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            fill = _FLAGGED_FILL if is_flagged_row else None
            _apply_cell_style(cell, font=_CELL_FONT, fill=fill,
                              alignment=left, border=_THIN_BORDER)

        # CSV row dict
        csv_rows.append(dict(zip(all_headers, row_data)))

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    ws.column_dimensions[get_column_letter(1)].width = 22   # Participant_ID
    ws.column_dimensions[get_column_letter(2)].width = 22   # Name
    ws.column_dimensions[get_column_letter(3)].width = 14   # Date
    for col_idx in range(4, len(question_cols) + 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions[get_column_letter(n_cols)].width = 28   # Flagged

    # Freeze panes so headers stay visible while scrolling
    ws.freeze_panes = "A4"

    return pd.DataFrame(csv_rows, columns=all_headers)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class OutputWriter:
    """Writes Excel + CSV output from processed ParticipantRecord lists."""

    def __init__(self, study_name: str = STUDY_NAME) -> None:
        self.study_name = study_name
        OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    def write(
        self,
        pre_results: List[dict],
        post_results: List[dict],
    ) -> Dict[str, Path]:
        """
        Generate the Excel workbook and CSV files.

        Args:
            pre_results:  List of ParticipantRecord dicts for PRE surveys.
            post_results: List of ParticipantRecord dicts for POST surveys.

        Returns:
            dict with keys 'excel', 'pre_csv', 'post_csv' mapping to Paths.
        """
        safe_name = re.sub(r"[^\w\-]", "_", self.study_name)
        excel_path = OUTPUT_FOLDER / f"{safe_name}_results.xlsx"
        pre_csv_path = OUTPUT_FOLDER / f"{safe_name}_PRE_results.csv"
        post_csv_path = OUTPUT_FOLDER / f"{safe_name}_POST_results.csv"

        # Collect question columns per survey type
        pre_q_cols  = _collect_question_columns(pre_results)
        post_q_cols = _collect_question_columns(post_results)

        # Build workbook
        wb = Workbook()
        # Remove the default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        output_paths: Dict[str, Path] = {"excel": excel_path}

        # PRE sheet
        pre_df = _write_sheet(
            wb,
            sheet_title="PRE",
            records=pre_results,
            question_cols=pre_q_cols,
            survey_label="PRE Survey Results",
        )

        # POST sheet
        post_df = _write_sheet(
            wb,
            sheet_title="POST",
            records=post_results,
            question_cols=post_q_cols,
            survey_label="POST Survey Results",
        )

        # If both sheets are empty, at least create a placeholder
        if not wb.sheetnames:
            ws = wb.create_sheet("No Data")
            ws["A1"] = "No survey records were processed."

        # Save workbook
        wb.save(str(excel_path))
        logger.info(f"Excel saved: {excel_path}")

        # Export CSVs
        if pre_df is not None:
            pre_df.to_csv(str(pre_csv_path), index=False, encoding="utf-8-sig")
            logger.info(f"PRE CSV saved: {pre_csv_path}")
            output_paths["pre_csv"] = pre_csv_path

        if post_df is not None:
            post_df.to_csv(str(post_csv_path), index=False, encoding="utf-8-sig")
            logger.info(f"POST CSV saved: {post_csv_path}")
            output_paths["post_csv"] = post_csv_path

        # Write a brief summary sheet
        self._write_summary_sheet(wb, pre_results, post_results, excel_path)

        return output_paths

    @staticmethod
    def _write_summary_sheet(
        wb: Workbook,
        pre_results: List[dict],
        post_results: List[dict],
        excel_path: Path,
    ) -> None:
        """Append a Summary sheet and re-save the workbook."""
        ws = wb.create_sheet(title="Summary", index=0)

        center = Alignment(horizontal="center", vertical="center")
        left   = Alignment(horizontal="left",   vertical="center")

        def _hdr(row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font = _HEADER_FONT
            c.fill = _HEADER_ROW_FILL
            c.alignment = center
            c.border = _THIN_BORDER

        def _val(row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font = _CELL_FONT
            c.alignment = left
            c.border = _THIN_BORDER

        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = f"Processing Summary — {STUDY_NAME}"
        title_cell.font = _WHITE_FONT
        title_cell.fill = _STUDY_ROW_FILL
        title_cell.alignment = center
        ws.row_dimensions[1].height = 22

        _hdr(2, 1, "Metric")
        _hdr(2, 2, "PRE")
        _hdr(2, 3, "POST")

        def _counts(results):
            total   = len(results)
            success = sum(1 for r in results if r.get("status") == "success")
            flagged = sum(1 for r in results if r.get("status") == "flagged")
            errors  = sum(1 for r in results if r.get("status") == "error")
            return total, success, flagged, errors

        pt, ps, pf, pe = _counts(pre_results)
        ot, os_, of_, oe = _counts(post_results)

        rows = [
            ("Total records",        pt,  ot),
            ("Successful",           ps,  os_),
            ("Flagged for review",   pf,  of_),
            ("Errors",               pe,  oe),
            ("Generated (UTC)",      datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"), ""),
        ]
        for r_idx, (metric, pre_val, post_val) in enumerate(rows, start=3):
            _val(r_idx, 1, metric)
            _val(r_idx, 2, pre_val)
            _val(r_idx, 3, post_val)

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14

        wb.save(str(excel_path))
